from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import csv

BASE = Path(__file__).resolve().parent
html = BASE / 'calculadora-agentes-beglobal.html'
text = html.read_text(encoding='utf-8')
repls = {
    '<div class="field full"><label>Pago único de implementación</label><input id="setupExtra" type="number" value="75000" /></div>': '<div class="field full"><label>Setup estratégico adicional</label><input id="setupExtra" type="number" value="75000" /></div>',
    'Este pago único se suma al costo de implementación. No se considera gasto operativo mensual.': 'Este setup estratégico adicional ($75,000 MXN) se suma al costo de implementación. No se considera gasto operativo mensual.',
    'Además, el escenario contempla un pago único de implementación de ${money(num(\'setupExtra\'))} MXN. Este pago no es mensual; se suma una sola vez al costo de implementación.': 'Además, el escenario contempla un setup estratégico adicional de ${money(num(\'setupExtra\'))} MXN. Este monto no es mensual; se suma una sola vez al costo de implementación.',
}
for old, new in repls.items():
    text = text.replace(old, new)
html.write_text(text, encoding='utf-8')

csv_rows = [
    ['Sección','Concepto','Valor MXN / %','Notas'],
    ['Aportación','Aportación inicial mínima',35000,'Base desde $35,000 MXN'],
    ['Aportación','Crecimiento',50000,'Comisión sugerida 15%'],
    ['Aportación','Crecimiento Plus',75000,'Comisión sugerida 20%'],
    ['Aportación','Premium',100000,'Comisión sugerida 25% o negociada'],
    ['Implementación','Horas internas',60,'Editable'],
    ['Implementación','Costo hora interno',350,'Editable'],
    ['Implementación','Integraciones',5000,'Editable'],
    ['Implementación','QA / capacitación',4000,'Editable'],
    ['Implementación','Setup estratégico adicional',75000,'Pago único; no gasto mensual'],
    ['Implementación','Costo implementación base',105000,'60*350 + 5000 + 4000 + 75000'],
    ['Operación mensual','IA / tokens',2500,'Editable'],
    ['Operación mensual','Hosting / herramientas',1800,'Editable'],
    ['Operación mensual','WhatsApp / CRM / Apps',2500,'Editable'],
    ['Operación mensual','Soporte humano',8000,'Editable'],
    ['Operación mensual','Salario ingenieros',25000,'Editable'],
    ['Operación mensual','Salario mercadólogos',18000,'Editable'],
    ['Venta','Precio venta',2500,'Editable'],
    ['Venta','Costo producto/proveedor',1500,'Editable'],
    ['Venta','Pasarela %',4,'Editable'],
    ['Venta','Envío/subsidio',150,'Editable'],
    ['Venta','Costo adquisición',120,'Editable'],
    ['Venta','Reserva riesgo %',5,'Editable'],
    ['Venta','Ventas mensuales estimadas',40,'Editable'],
    ['Venta','Mensualidad cobrada al cliente',12000,'Editable'],
]
with (BASE/'calculadora-agentes-beglobal-import.csv').open('w', newline='', encoding='utf-8') as f:
    csv.writer(f).writerows(csv_rows)

wb = Workbook()
ws = wb.active
ws.title = 'Resumen ejecutivo'
rows = [
    ['Calculadora Be Global Smart Agent','',''],
    ['Aportación inicial mínima',35000,'MXN'],
    ['Setup estratégico adicional',75000,'MXN, pago único'],
    ['Costo implementación base','=Configuración!B4*Configuración!B5+Configuración!B6+Configuración!B7+Configuración!B8','MXN'],
    ['Gasto operativo mensual','=SUM(Operación!B2:B7)','MXN'],
    ['Utilidad neta por venta','=Ventas!B2-Ventas!B3-(Ventas!B2*Ventas!B4)-Ventas!B5-Ventas!B6-(Ventas!B2*Ventas!B7)','MXN'],
    ['Comisión Allan/Be Global por venta','=MAX(0,B6*Comisiones!B2)','MXN'],
    ['Utilidad mensual estimada','=B6*Ventas!B8+Ventas!B9-B5','MXN'],
    ['Ventas para cubrir operación','=ROUNDUP(B5/B6,0)','ventas'],
    ['Meses para recuperar implementación','=IF(B8>0,B4/B8,"No rentable")','meses'],
]
for r in rows: ws.append(r)

config = wb.create_sheet('Configuración')
for r in [['Concepto','Valor','Notas'],['Aportación inicial',35000,'Desde $35,000 MXN'],['Comisión sugerida',0.10,'Base: 10%; crecimiento 15%; plus 20%; premium 25%'],['Horas internas',60,'Editable'],['Costo hora interno',350,'Editable'],['Integraciones',5000,'Editable'],['QA / capacitación',4000,'Editable'],['Setup estratégico adicional',75000,'Pago único; no gasto mensual']]: config.append(r)
ops = wb.create_sheet('Operación')
for r in [['Concepto','Valor MXN','Notas'],['IA / tokens',2500,'Editable'],['Hosting / herramientas',1800,'Editable'],['WhatsApp / CRM / Apps',2500,'Editable'],['Soporte humano',8000,'Editable'],['Salario ingenieros',25000,'Editable'],['Salario mercadólogos',18000,'Editable']]: ops.append(r)
ventas = wb.create_sheet('Ventas')
for r in [['Concepto','Valor','Notas'],['Precio venta',2500,'MXN'],['Costo producto/proveedor',1500,'MXN'],['Pasarela %',0.04,'Porcentaje'],['Envío/subsidio',150,'MXN'],['Costo adquisición',120,'MXN'],['Reserva riesgo %',0.05,'Porcentaje'],['Ventas mensuales estimadas',40,'Unidades'],['Mensualidad cobrada al cliente',12000,'MXN']]: ventas.append(r)
com = wb.create_sheet('Comisiones')
for r in [['Concepto','Valor','Notas'],['Comisión Allan/Be Global %',0.10,'Sobre utilidad neta, no venta bruta'],['Base',0.10,'Aportación $35,000'],['Crecimiento',0.15,'Aportación $50,000'],['Crecimiento Plus',0.20,'Aportación $75,000'],['Premium',0.25,'Aportación $100,000 o negociada']]: com.append(r)
esc = wb.create_sheet('Escenarios')
for r in [['Escenario','Ventas','Utilidad mensual','Comisión mensual','Lectura'],['Conservador','=Ventas!B8*0.6','=Resumen ejecutivo!B6*B2+Ventas!B9-Resumen ejecutivo!B5','=Resumen ejecutivo!B7*B2','=IF(C2>0,"Rentable","No rentable")'],['Base','=Ventas!B8','=Resumen ejecutivo!B8','=Resumen ejecutivo!B7*B3','=IF(C3>0,"Rentable","No rentable")'],['Agresivo','=Ventas!B8*1.6','=Resumen ejecutivo!B6*B4+Ventas!B9-Resumen ejecutivo!B5','=Resumen ejecutivo!B7*B4','=IF(C4>0,"Rentable","No rentable")']]: esc.append(r)
for sheet in wb.worksheets:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='172554')
    for col in range(1,4):
        sheet.column_dimensions[get_column_letter(col)].width = 30
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
wb.save(BASE/'calculadora-agentes-beglobal.xlsx')

proposal = '''PROPUESTA COMERCIAL — BE GLOBAL SMART AGENT

Proyecto: Agente Be Global

1. Objetivo
Implementar un Be Global Smart Agent para apoyar ventas, atención, catálogo u operación ecommerce con diagnóstico, configuración, base de conocimiento, pruebas y seguimiento operativo.

2. Aportación inicial
Aportación inicial desde $35,000 MXN.

3. Costos de implementación
Costo de implementación base actualizado: $105,000 MXN.
Incluye horas internas, integraciones, QA/capacitación y setup estratégico adicional de $75,000 MXN.

Nota: el setup estratégico adicional es pago único; no se considera gasto operativo mensual.

4. Operación mensual
Debe contemplar IA/tokens, hosting/herramientas, WhatsApp/CRM/apps, soporte humano, salario de ingenieros y salario de mercadólogos.

5. Comisión de alianza estratégica
La comisión Allan/Be Global se calcula sobre utilidad neta, no sobre venta bruta.
Regla sugerida: a mayor aportación inicial, mayor porcentaje de comisión.

6. Condiciones importantes
Esta herramienta es de planeación. No garantiza ventas ni ingresos. Antes de presentar propuesta final se deben validar alcance, costos reales, contratos, impuestos, pagos, soporte y operación.

Siguiente paso: definir si el setup estratégico adicional de $75,000 MXN será fijo para todos los niveles o variable según la aportación inicial.
'''
(BASE/'propuesta-comercial-beglobal-smart-agent.txt').write_text(proposal, encoding='utf-8')
print('Regenerados: HTML, XLSX, CSV y propuesta TXT')
